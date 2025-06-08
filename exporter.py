# exporter.py
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import Rule, CellIsRule
from openpyxl.utils import get_column_letter

def _format_worksheet(worksheet):
    """Applies conditional formatting and adjusts column widths for a worksheet."""
    # (This is your original format_worksheet function, moved here)
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Find the 'گذرانده شده' column
    header = [cell.value for cell in worksheet[1]]
    try:
        status_col_letter = get_column_letter(header.index('گذرانده شده') + 1)
    except ValueError:
        return # Column not found, can't format

    # Add formatting rules
    worksheet.conditional_formatting.add(
        f'{status_col_letter}2:{status_col_letter}{worksheet.max_row}',
        CellIsRule(operator='equal', formula=['"بله"'], fill=green_fill)
    )
    worksheet.conditional_formatting.add(
        f'{status_col_letter}2:{status_col_letter}{worksheet.max_row}',
        CellIsRule(operator='equal', formula=['"خیر"'], fill=red_fill)
    )

    # Auto-adjust column widths
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width


class Exporter:
    """Handles exporting data to formatted Excel files."""

    def __init__(self, training_analyzer):
        self.analyzer = training_analyzer

    def export_single_dealer(self, dealer_name, filename):
        """Exports a single dealer's training analysis to an Excel file."""
        df = self.analyzer.generate_dealer_export_df(dealer_name)
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            sheet_name = dealer_name.split(' - ')[-1][:30] if ' - ' in dealer_name else dealer_name[:30]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            _format_worksheet(worksheet)

    def export_all_dealers(self, dealer_names, filename):
        """Exports all dealers' training analysis to a single Excel file, each on its own sheet."""
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            for dealer_name in dealer_names:
                df = self.analyzer.generate_dealer_export_df(dealer_name)
                sheet_name = dealer_name.split(' - ')[-1][:30] if ' - ' in dealer_name else dealer_name[:30]
                
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                _format_worksheet(worksheet)